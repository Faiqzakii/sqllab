import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pyarrow.parquet as parquet

from sql_lab_extractor.artifacts import RunStore
from sql_lab_extractor.finalize import finalize_run


class FinalizationTests(unittest.TestCase):
    def _store_with_pages(self, root: Path):
        store = RunStore.create(root, "SELECT id, name FROM sample ORDER BY id")
        later = store.write_page(1000, [{"id": 3, "name": "third"}])
        first = store.write_page(0, [{"id": 1, "name": "first"}, {"id": 2, "name": "second"}])
        return store, (later, first)

    def test_writes_offset_ordered_parquet_before_excel(self):
        with tempfile.TemporaryDirectory() as temporary:
            store, records = self._store_with_pages(Path(temporary))

            result = finalize_run(store, records, ("parquet", "xlsx"))

            self.assertEqual(result.status, "completed")
            self.assertEqual(result.rows, 3)
            self.assertEqual(result.columns, ("id", "name"))
            self.assertTrue(result.parquet_path.is_file())
            self.assertTrue(result.excel_path.is_file())
            self.assertEqual(
                parquet.read_table(result.parquet_path).to_pylist(),
                [{"id": "1", "name": "first"}, {"id": "2", "name": "second"}, {"id": "3", "name": "third"}],
            )

    def test_rejects_schema_drift_between_valid_pages(self):
        with tempfile.TemporaryDirectory() as temporary:
            store = RunStore.create(Path(temporary), "SELECT id FROM sample ORDER BY id")
            first = store.write_page(0, [{"id": 1, "name": "first"}])
            second = store.write_page(1000, [{"id": 2, "other": "second"}])

            with self.assertRaisesRegex(ValueError, "Schema drift"):
                finalize_run(store, (first, second), ("parquet",))
            self.assertFalse((store.run_dir / "result.parquet").exists())

    def test_rejects_record_not_matching_manifest(self):
        with tempfile.TemporaryDirectory() as temporary:
            store, (record, _) = self._store_with_pages(Path(temporary))

            forged = type(record)(record.offset, record.rows + 1, record.columns, record.checksum)

            with self.assertRaisesRegex(ValueError, "manifest"):
                finalize_run(store, (forged,), ("parquet",))

    def test_orders_generator_records_without_materializing_a_sorted_tuple(self):
        with tempfile.TemporaryDirectory() as temporary:
            store, records = self._store_with_pages(Path(temporary))

            result = finalize_run(store, (record for record in records), ("parquet",))

            self.assertEqual(
                parquet.read_table(result.parquet_path).column("id").to_pylist(),
                ["1", "2", "3"],
            )
    def test_closes_parquet_file_after_excel_generation(self):
        with tempfile.TemporaryDirectory() as temporary:
            store, records = self._store_with_pages(Path(temporary))
            parquet_path = finalize_run(store, records, ("parquet",)).parquet_path
            original_parquet_file = parquet.ParquetFile
            opened_files = []

            class TrackingParquetFile:
                def __init__(self, path):
                    self._file = original_parquet_file(path)
                    self.schema_arrow = self._file.schema_arrow
                    opened_files.append(self)

                def __enter__(self):
                    return self

                def __exit__(self, exception_type, exception, traceback):
                    self._file.close()

                def iter_batches(self, **kwargs):
                    return self._file.iter_batches(**kwargs)

            with patch("sql_lab_extractor.finalize.parquet.ParquetFile", TrackingParquetFile):
                from sql_lab_extractor.finalize import _write_excel

                _write_excel(parquet_path, store.run_dir / "result.xlsx", 2)

            self.assertTrue(opened_files)
            self.assertTrue(opened_files[0]._file.closed)


    def test_splits_excel_sheets_at_injected_row_limit(self):
        with tempfile.TemporaryDirectory() as temporary:
            store, records = self._store_with_pages(Path(temporary))

            result = finalize_run(store, records, ("xlsx",), excel_row_limit=2)

            from openpyxl import load_workbook
            workbook = load_workbook(result.excel_path, read_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Sheet1", "Sheet2"])
                self.assertEqual(list(workbook["Sheet1"].values), [("id", "name"), ("1", "first"), ("2", "second")])
                self.assertEqual(list(workbook["Sheet2"].values), [("id", "name"), ("3", "third")])
            finally:
                workbook.close()

    def test_preserves_parquet_when_excel_generation_fails(self):
        with tempfile.TemporaryDirectory() as temporary:
            store, records = self._store_with_pages(Path(temporary))

            with patch("sql_lab_extractor.finalize._write_excel", side_effect=OSError("locked")):
                result = finalize_run(store, records, ("parquet", "xlsx"))

            self.assertEqual(result.status, "completed_with_excel_error")
            self.assertTrue(result.parquet_path.is_file())
            self.assertIsNone(result.excel_path)


if __name__ == "__main__":
    unittest.main()
