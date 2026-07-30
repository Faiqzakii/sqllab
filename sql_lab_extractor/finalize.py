from __future__ import annotations

import hashlib
import os
import tempfile
from itertools import chain
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pyarrow as pa
import pyarrow.parquet as parquet
from openpyxl import Workbook

from .artifacts import PageRecord, RunStore

DEFAULT_EXCEL_ROW_LIMIT = 1_048_575
PARQUET_BATCH_SIZE = 10_000


@dataclass(frozen=True)
class FinalizeResult:
    status: str
    parquet_path: Path
    excel_path: Path | None
    rows: int
    columns: tuple[str, ...]


def finalize_run(
    run_store: RunStore,
    records: Iterable[PageRecord],
    formats: Iterable[str],
    *,
    excel_row_limit: int = DEFAULT_EXCEL_ROW_LIMIT,
) -> FinalizeResult:
    """Validate ordered page CSVs, write Parquet, then optionally Excel."""
    selected_formats = tuple(formats)
    _validate_formats(selected_formats, excel_row_limit)
    records_by_offset: dict[int, PageRecord] = {}
    for record in records:
        if record.offset in records_by_offset:
            raise ValueError("Offset halaman duplikat")
        records_by_offset[record.offset] = record
    columns = _validate_records(run_store, records_by_offset.values())
    parquet_path = run_store.run_dir / "result.parquet"
    rows = _write_parquet(run_store, _iter_ordered_records(records_by_offset), columns, parquet_path)
    if "xlsx" not in selected_formats:
        return FinalizeResult("completed", parquet_path, None, rows, columns)
    excel_path = run_store.run_dir / "result.xlsx"
    try:
        _write_excel(parquet_path, excel_path, excel_row_limit)
    except Exception:
        return FinalizeResult("completed_with_excel_error", parquet_path, None, rows, columns)
    return FinalizeResult("completed", parquet_path, excel_path, rows, columns)


def _validate_formats(formats: tuple[str, ...], excel_row_limit: int) -> None:
    if not formats or any(output_format not in {"parquet", "xlsx"} for output_format in formats):
        raise ValueError("Format akhir wajib parquet dan/atau xlsx")
    if excel_row_limit < 1:
        raise ValueError("Batas baris Excel wajib positif")


def _iter_ordered_records(records_by_offset: dict[int, PageRecord]) -> Iterable[PageRecord]:
    for offset in sorted(records_by_offset):
        yield records_by_offset[offset]


def _validate_records(run_store: RunStore, records: Iterable[PageRecord]) -> tuple[str, ...]:
    iterator = iter(records)
    try:
        first_record = next(iterator)
    except StopIteration:
        raise ValueError("Tidak ada halaman untuk difinalisasi") from None
    columns = first_record.columns
    if not columns:
        raise ValueError("Halaman tanpa kolom tidak dapat difinalisasi")
    for record in chain((first_record,), iterator):
        manifest_record = run_store._manifest_page_record(record.offset)
        if manifest_record != record:
            raise ValueError("Rekaman halaman tidak cocok dengan manifest")
        rows = run_store.load_valid_page(record.offset)
        if rows is None:
            raise ValueError("Halaman tidak valid")
        if len(rows) != record.rows:
            raise ValueError("Jumlah baris halaman tidak valid")
        if record.columns != columns:
            raise ValueError("Schema drift terdeteksi antar halaman")
    return columns


def _checksum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def _write_parquet(
    run_store: RunStore,
    records: Iterable[PageRecord],
    columns: tuple[str, ...],
    target: Path,
) -> int:
    schema = pa.schema([pa.field(column, pa.string()) for column in columns])
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{target.name}.", suffix=".partial", dir=target.parent)
    os.close(descriptor)
    temporary = Path(temporary_name)
    total_rows = 0
    try:
        with parquet.ParquetWriter(temporary, schema) as writer:
            for record in records:
                rows = run_store.load_valid_page(record.offset)
                if rows is None:
                    raise ValueError("Halaman tidak valid")
                batch: list[dict[str, str | None]] = []
                for row in rows:
                    if set(row) != set(columns):
                        raise ValueError("Schema drift terdeteksi antar halaman")
                    batch.append({column: None if row[column] is None else str(row[column]) for column in columns})
                    if len(batch) == PARQUET_BATCH_SIZE:
                        writer.write_table(pa.Table.from_pylist(batch, schema=schema))
                        total_rows += len(batch)
                        batch.clear()
                if batch:
                    writer.write_table(pa.Table.from_pylist(batch, schema=schema))
                    total_rows += len(batch)
        os.replace(temporary, target)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise
    return total_rows


def _write_excel(parquet_path: Path, target: Path, row_limit: int) -> None:
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{target.stem}.", suffix=".xlsx", dir=target.parent)
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        workbook = Workbook(write_only=True)
        with parquet.ParquetFile(parquet_path) as parquet_file:
            columns = tuple(parquet_file.schema_arrow.names)
            sheet_number = 0
            rows_in_sheet = row_limit
            worksheet = None
            for batch in parquet_file.iter_batches(batch_size=PARQUET_BATCH_SIZE):
                for row in batch.to_pylist():
                    if rows_in_sheet >= row_limit:
                        sheet_number += 1
                        worksheet = workbook.create_sheet(f"Sheet{sheet_number}")
                        worksheet.append(columns)
                        rows_in_sheet = 0
                    worksheet.append([row[column] for column in columns])
                    rows_in_sheet += 1
            if worksheet is None:
                worksheet = workbook.create_sheet("Sheet1")
                worksheet.append(columns)
        workbook.save(temporary)
        os.replace(temporary, target)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise
